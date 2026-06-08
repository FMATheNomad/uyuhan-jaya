import io
import json
import re
from typing import AsyncGenerator, Optional
from app.core.config import settings


async def call_ai(messages: list[dict], model: str = None) -> str:
    api_key = settings.deepseek_api_key or settings.openrouter_api_key
    model = model or settings.deepseek_model or settings.openrouter_model
    base_url = settings.ai_base_url

    if not api_key:
        return "AI API key not configured."

    import httpx

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    if "openrouter" in base_url:
        headers["HTTP-Referer"] = "https://minicrane.up.railway.app"
        headers["X-Title"] = "MiniCrane"

    async with httpx.AsyncClient(timeout=60.0) as client:
        response = await client.post(
            f"{base_url}/chat/completions",
            headers=headers,
            json={
                "model": model,
                "messages": messages,
                "stream": False,
                "max_tokens": 4096,
                "temperature": 0.3,
            },
        )
        if response.status_code != 200:
            return f"Error: AI service returned status {response.status_code}"
        data = response.json()
        return data.get("choices", [{}])[0].get("message", {}).get("content", "")


RAB_SYSTEM_PROMPT = """Anda adalah estimator konstruksi Indonesia yang ahli membuat RAB (Rencana Anggaran Biaya).

Tugas Anda adalah membuat RAB dalam format JSON yang VALID dan TERSTRUKTUR.

Output Anda HARUS hanya berupa JSON valid (tanpa markdown, tanpa ```, tanpa penjelasan apapun) dengan format:

{
  "project_name": "Nama Proyek",
  "description": "Deskripsi singkat proyek",
  "location": "Estimasi lokasi",
  "total_area_m2": 0,
  "price_per_m2": 0,
  "sections": [
    {
      "name": "NAMA PEKERJAAN",
      "items": [
        {
          "description": "Uraian pekerjaan",
          "volume": 0,
          "unit": "m³/m²/buah/LS/dll",
          "unit_price": 0,
          "total": 0
        }
      ],
      "subtotal": 0
    }
  ],
  "total_direct_cost": 0,
  "contingency_pct": 10,
  "contingency": 0,
  "grand_total": 0
}

Aturan:
- Gunakan satuan standar konstruksi Indonesia (m³, m², m', buah, sak, kg, LS)
- Gunakan harga wajar untuk proyek kelas menengah
- Sections minimal: Pekerjaan Persiapan, Pondasi, Struktur, Dinding, Atap, Finishing
- Setiap section HARUS punya minimal 2 item
- Hitung total = volume * unit_price
- Subtotal = jumlah semua total item di section
- Contingency = 10% dari total_direct_cost
- grand_total = total_direct_cost + contingency
- Gunakan angka tanpa titik/koma sebagai pemisah ribuan (contoh: 500000 bukan 500,000)
- Output HARUS hanya JSON, tidak ada teks lain"""


async def generate_rab_ai(project_description: str) -> dict:
    prompt = f"""Buat RAB untuk proyek dengan deskripsi berikut:
{project_description}

Output dalam format JSON sesuai instruksi sebelumnya."""
    
    messages = [
        {"role": "system", "content": RAB_SYSTEM_PROMPT},
        {"role": "user", "content": prompt},
    ]
    
    result = await call_ai(messages)
    
    try:
        json_match = re.search(r'\{.*\}', result, re.DOTALL)
        if json_match:
            parsed = json.loads(json_match.group())
            return parsed
        parsed = json.loads(result)
        return parsed
    except (json.JSONDecodeError, ValueError):
        return {
            "raw_text": result,
            "parse_error": True,
            "message": "AI gagal menghasilkan format terstruktur. Tampilkan teks mentah ini sebagai fallback."
        }


async def generate_rab_excel(rab_data: dict) -> Optional[bytes]:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RAB"

    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    total_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    rp_format = '#,##0'

    ws.merge_cells('A1:E1')
    ws['A1'] = f"RENCANA ANGGARAN BIAYA (RAB)"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:E2')
    ws['A2'] = rab_data.get('project_name', '')
    ws['A2'].font = Font(size=12)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:E3')
    desc = rab_data.get('description', '')
    ws['A3'] = desc[:100] if desc else ''
    ws['A3'].font = Font(size=10, color="666666")
    ws['A3'].alignment = Alignment(horizontal='center')

    if rab_data.get('total_area_m2'):
        ws.merge_cells('A4:E4')
        ws['A4'] = f"Luas Bangunan: {rab_data['total_area_m2']} m²  |  Harga/m²: Rp {rab_data.get('price_per_m2', 0):,}"
        ws['A4'].font = Font(size=10)

    start_row = 6
    headers = ['No', 'Uraian Pekerjaan', 'Volume', 'Satuan', 'Harga Satuan (Rp)', 'Total (Rp)']
    col_widths = [5, 50, 10, 8, 18, 18]

    row = start_row
    ws.cell(row=row, column=1, value='No').font = header_font
    for i, h in enumerate(headers[1:], 2):
        ws.cell(row=row, column=i, value=h).font = header_font
    for i in range(1, 7):
        ws.cell(row=row, column=i).border = thin_border
        ws.cell(row=row, column=i).alignment = Alignment(horizontal='center', wrap_text=True)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row += 1
    item_no = 1

    for section in rab_data.get('sections', []):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=section.get('name', ''))
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal='left')
        for i in range(1, 7):
            ws.cell(row=row, column=i).border = thin_border
        row += 1

        for item in section.get('items', []):
            ws.cell(row=row, column=1, value=item_no).border = thin_border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=item.get('description', '')).border = thin_border
            ws.cell(row=row, column=3, value=item.get('volume', 0)).border = thin_border
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4, value=item.get('unit', '')).border = thin_border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')

            c5 = ws.cell(row=row, column=5, value=int(item.get('unit_price', 0)))
            c5.border = thin_border
            c5.number_format = rp_format
            c5.alignment = Alignment(horizontal='right')

            c6 = ws.cell(row=row, column=6, value=int(item.get('total', 0)))
            c6.border = thin_border
            c6.number_format = rp_format
            c6.alignment = Alignment(horizontal='right')

            item_no += 1
            row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value='Subtotal').font = total_font
        ws.cell(row=row, column=1).border = thin_border
        for i in range(2, 5):
            ws.cell(row=row, column=i).border = thin_border
        c5 = ws.cell(row=row, column=5, value=int(section.get('subtotal', 0)))
        c5.font = total_font
        c5.border = thin_border
        c5.number_format = rp_format
        c5.alignment = Alignment(horizontal='right')
        c6 = ws.cell(row=row, column=6, value=int(section.get('subtotal', 0)))
        c6.font = total_font
        c6.border = thin_border
        c6.number_format = rp_format
        c6.alignment = Alignment(horizontal='right')

        row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value='TOTAL BIAYA LANGSUNG').font = total_font
    ws.cell(row=row, column=1).border = thin_border
    for i in range(2, 5):
        ws.cell(row=row, column=i).border = thin_border
    c5 = ws.cell(row=row, column=5)
    c5.border = thin_border
    c6 = ws.cell(row=row, column=6, value=int(rab_data.get('total_direct_cost', 0)))
    c6.font = total_font
    c6.border = thin_border
    c6.number_format = rp_format
    c6.alignment = Alignment(horizontal='right')
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=f"Biaya Tak Terduga ({rab_data.get('contingency_pct', 10)}%)").font = total_font
    ws.cell(row=row, column=1).border = thin_border
    for i in range(2, 5):
        ws.cell(row=row, column=i).border = thin_border
    c5 = ws.cell(row=row, column=5)
    c5.border = thin_border
    c6 = ws.cell(row=row, column=6, value=int(rab_data.get('contingency', 0)))
    c6.font = total_font
    c6.border = thin_border
    c6.number_format = rp_format
    c6.alignment = Alignment(horizontal='right')
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value='TOTAL ANGGARAN').font = Font(bold=True, size=13)
    ws.cell(row=row, column=1).border = thin_border
    for i in range(2, 5):
        ws.cell(row=row, column=i).border = thin_border
    c5 = ws.cell(row=row, column=5)
    c5.border = thin_border
    c6 = ws.cell(row=row, column=6, value=int(rab_data.get('grand_total', 0)))
    c6.font = Font(bold=True, size=13)
    c6.border = thin_border
    c6.number_format = rp_format
    c6.alignment = Alignment(horizontal='right')

    row += 2
    if rab_data.get('price_per_m2'):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=f"Harga per meter persegi: Rp {rab_data['price_per_m2']:,} / m²").font = Font(italic=True, color="666666")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


async def analyze_progress_with_ai(project_data: dict) -> str:
    prompt = f"""Analisis progres proyek konstruksi berikut:

Data proyek: {json.dumps(project_data, indent=2, default=str)}

Berikan:
1. Evaluasi progres saat ini
2. Potensi keterlambatan
3. Rekomendasi actionable
Gunakan bahasa Indonesia."""
    messages = [
        {"role": "system", "content": "Anda adalah konsultan manajemen konstruksi."},
        {"role": "user", "content": prompt},
    ]
    return await call_ai(messages)


async def predict_cashflow(project_data: dict) -> str:
    prompt = f"""Analisis cash flow proyek konstruksi berikut dan berikan prediksi 30/60/90 hari ke depan:

Data proyek: {json.dumps(project_data, indent=2, default=str)}

Berikan:
1. Prediksi cash flow per bulan
2. Peringatan jika ada potensi defisit
3. Rekomendasi antisipasi
Gunakan bahasa Indonesia."""
    messages = [
        {"role": "system", "content": "Anda adalah financial analyst spesialis konstruksi."},
        {"role": "user", "content": prompt},
    ]
    return await call_ai(messages)
